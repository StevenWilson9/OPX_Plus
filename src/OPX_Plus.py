import openpyxl, glob, webbrowser, warnings, csv, datetime, os, time
from copy import copy
from openpyxl.formula.translate import Translator


def to_unix_time(year=datetime.datetime.now().year, month=datetime.datetime.now().month,
                 day=datetime.datetime.now().day, hour=datetime.datetime.now().hour,
                 minute=datetime.datetime.now().minute, second=datetime.datetime.now().second, add_3_zeros=False,
                 subtract=datetime.timedelta(0)):
    date_time = datetime.datetime(year, month, day, hour, minute, second)
    date_time = date_time - subtract
    unix_time = int(time.mktime(date_time.timetuple()))
    og_unix_time = unix_time
    if add_3_zeros:
        unix_time = unix_time * 1000
    print(f"Converted: {date_time} -> {unix_time}")
    print("\t\t\t"+ datetime.datetime.utcfromtimestamp(og_unix_time).strftime('%Y-%m-%d %H:%M:%S') + " GMT")
    return unix_time
    # to_unix_time()
    # to_unix_time(add_3_zeros=True)
    # to_unix_time(second=0, minute=0)
    # subtract = datetime.timedelta(days=0, seconds=0)


def date_ranger():
    yesterday_date = datetime.date.today() - datetime.timedelta(days=1)

    yesterday_day_of_week = yesterday_date.strftime("%w")
    if yesterday_day_of_week == 0:
        days_back = 3
    elif yesterday_day_of_week == 1:
        days_back = 4
    else:
        days_back = 2

    period_start_date = datetime.date.today() - datetime.timedelta(days=days_back)

    if period_start_date.strftime("%b") == yesterday_date.strftime("%b"):
        combined_date = f'{period_start_date.strftime("%b %d")}-{yesterday_date.strftime("%d")}'
    else:
        combined_date = f'{period_start_date.strftime("%b %d")}-{yesterday_date.strftime("%b %d")}'

    return combined_date + f' {yesterday_date.strftime("%Y")}'


def get_file_path(file_path, name_search, *if_missing_urls):
    full_file_path = glob.glob(f'{file_path}{name_search}')

    if len(full_file_path) == 0:
        error_message = True
        for url in if_missing_urls:
            if url.startswith("http") or url.startswith("file:/"):
                webbrowser.open(url)
            else:
                error_message = False
                print(url)
        if error_message:
            print(f'MISSING: {name_search} file in {file_path}.')
        return "not found"

    elif len(full_file_path) > 1:
        print("------------------")
        print(f'Multiple files fitting {name_search} search, found in {file_path}.')
        for path in full_file_path:
            print(f"- {path}")
        print("Narrow the search parameters, or delete all but one file, and rerun the script")
        print("------------------")
        return "not found"

    file_path = full_file_path[0]
    extension = os.path.splitext(file_path)[1]
    if extension not in ['.csv', '.xlsx']:
        print(f"ERROR: File extension {extension} is not supported. Only .csv and .xlsx files are supported.")
        return "not found"

    print(f"Found: {file_path.split('/')[-1]}")
    return file_path


def check_mandatory_files(f_list):
    if "not found" in f_list:
        print("------------------")
        quit()
    else:
        print("All mandatory files found")
        print("------------------")


def open_template(template_path: object, *date_cells: str) -> object:
    wb = openpyxl.load_workbook(template_path)
    ws = wb[wb.sheetnames[0]]
    print(f'Opened "{ws.title}" from "{template_path.split("/")[-1]}"')
    today_date = datetime.date.today()

    for cell in date_cells:
        ws[cell] = today_date
        print(f"cell {cell} updated with today's date")
    return wb, ws


def save_file(new_file_location, file_name, workbook, afterdatetext="", previousday=False):
    if afterdatetext != "":
        afterdatetext = " " + afterdatetext
    if previousday:
        new_file_path = f'{new_file_location}{file_name} {datetime.date.today() - datetime.timedelta(days=1)}{afterdatetext}.xlsx'
    else:
        new_file_path = f'{new_file_location}{file_name} {datetime.date.today()}{afterdatetext}.xlsx'
    workbook.save(new_file_path)
    print(f'saved at: {new_file_path}')
    print(f"Opening {new_file_path.split('/')[-1]}")
    print("------------------")
    new_file_path = f'"{new_file_path}"'
    os.system("open " + new_file_path)


def remove(file_path):
    if "*template*" in file_path:
        print(f'.py file has been directed to delete is attempting to delete {file_path.split("/")[-1]}')
        print(f'This has not been executed.')
    elif file_path == "not found":
        pass
    else:
        os.remove(file_path)
        print(f'DELETED: {file_path.split("/")[-1]}')


def open_vals_only_sheet(from_wb_path, sheet_id=0):
    with warnings.catch_warnings(record=True):
        warnings.simplefilter("always")
        values_only_wb = openpyxl.load_workbook(from_wb_path, data_only=True)
        if sheet_id is int or sheet_id == 0:
            values_only_sheet = values_only_wb[values_only_wb.sheetnames[sheet_id]]
        else:
            values_only_sheet = values_only_wb[sheet_id]
        if sheet_id == 0:
            print(f'Opened values version of "{from_wb_path.split("/")[-1]}"')
        else:
            print(f'Opened values version of "{values_only_sheet.title}" from "{from_wb_path.split("/")[-1]}"')
        return values_only_sheet


def paste_sheet_to_sheet(from_ws, to_ws, cell_range):
    rows = count_rows(from_ws)
    for row in openpyxl.utils.rows_from_range(f"{cell_range}{rows}"):
        for cell in row:
            to_ws[cell].value = from_ws[cell].value
    print(f'C&Ped VALUEs from "{from_ws.title}" to "{to_ws.title}" for {rows} rows')


def paste_csv_vals_to_sheet(csv_path, to_sheet, include_header=False):
    with open(csv_path, "r") as f:
        row_incrementer = 1
        reader = csv.reader(f)
        if not include_header:
            next(reader)
            row_incrementer += 1
        for row_index, row in enumerate(reader):
            for column_index, cell in enumerate(row):
                column_letter = num_to_excel_col((column_index + 1))
                s = cell
                try:
                    s = float(s)
                except ValueError:
                    pass
                to_sheet[f'{column_letter}{row_index + row_incrementer}'].value = s
    if include_header:
        print(f'C&Ped VALUE & HEADERs from {csv_path.split("/")[-1]} to "{to_sheet.title}"')
    else:
        print(f'C&Ped VALUEs from {csv_path.split("/")[-1]} to "{to_sheet.title}"')


def copy_over_and_down_formulas(from_ws, to_ws, formula_cells):
    for row in openpyxl.utils.rows_from_range(formula_cells):
        c_list = []
        for cell in row:
            c_list.append(cell.rstrip('1234567890'))
            to_ws[cell].value = from_ws[cell].value
    # drag down formulas
    rows = count_rows(to_ws)
    # print(f'c_list is {c_list}')
    for col in c_list:
        # print(f'col is {col}')
        for row in openpyxl.utils.rows_from_range(f"{col}3:{col}{rows}"):
            origin_cell = f"{col}2"
            target_cell = row[0]
            formula = to_ws[origin_cell].value
            to_ws[target_cell].value = Translator(formula, origin=origin_cell).translate_formula(target_cell)

            to_ws[target_cell].font = copy(to_ws[origin_cell].font)
            to_ws[target_cell].border = copy(to_ws[origin_cell].border)
            to_ws[target_cell].fill = copy(to_ws[origin_cell].fill)
            to_ws[target_cell].number_format = copy(to_ws[origin_cell].number_format)
            to_ws[target_cell].protection = copy(to_ws[origin_cell].protection)
            to_ws[target_cell].alignment = copy(to_ws[origin_cell].alignment)

    print(f'C&Ped FORUMULAs from "{from_ws.title}" to "{to_ws.title}" for {rows} rows')
    print("------------------")


def count_rows(worksheet):
    rows = 5
    cell = "B" + str(rows)

    while worksheet[cell].value is not None:
        cell = "B" + str(rows)
        rows += 1
    rows -= 2
    # print(f'{rows} rows in sheet "{worksheet.title}"')
    return rows


def num_to_excel_col(n):
    """
    Converts a column number to a Excel column alphabet
    eg 27 -> AA
    """

    if n < 1:
        raise ValueError("Number must be positive")
    result = ""
    while True:
        if n > 26:
            n, r = divmod(n - 1, 26)
            result = chr(r + ord('A')) + result
        else:
            return chr(n + ord('A') - 1) + result


def excel_col_to_num(a):
    """
    Converts a Excel column alphabet to a column number
    eg AA -> 27
    """

    if len(a) > 1:
        result = 0
        for l in [*a]:
            result += ord(l) - ord('A') + 1
            return result
    else:
        return ord(a) - ord('A') + 1


def paste_cells_to_cells(from_ws, to_ws, from_cell_range, offset=(0, 0)):
    x_offset, y_offset = offset
    for row in openpyxl.utils.rows_from_range(from_cell_range):
        for cell in row:
            col2 = num_to_excel_col(excel_col_to_num(cell.rstrip('1234567890')) + x_offset)
            row2 = int(cell.lstrip('abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ')) + y_offset
            cell2 = f"{col2}{row2}"

            to_ws[cell2] = from_ws[cell].value
    print(f'C&Ped VALUEs from "{from_ws.title}" to "{to_ws.title}"')


def worksheet_reset(sheet_nam, wb):
    ws = wb[sheet_nam]
    ws.auto_filter.ref = None

    row = 1
    while row <= ws.max_row:
        # unhide rows as ya go
        ws.row_dimensions[row].hidden = False
        row += 1
    ws.delete_rows(10, ws.max_row+10)

    # count columns
    column_num = 1
    cell = num_to_excel_col(column_num) + "1"

    while ws[cell].value is not None:
        cell = num_to_excel_col(column_num) + "1"
        column_num += 1
    column_num -= 2
    cell = num_to_excel_col(column_num) + "1"
    print(f'"{sheet_nam}" goes to column {ws[cell].value} ({num_to_excel_col(column_num)}. Is reset.)')

